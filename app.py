# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timedelta
import os, json, imaplib, email, shutil
from email.header import decode_header
from dotenv import load_dotenv
import anthropic
import fitz

load_dotenv(override=True)

app = Flask(__name__)
app.secret_key = 'gyomu_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gyomu.db'
app.config['UPLOAD_FOLDER'] = 'uploads'
db = SQLAlchemy(app)

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

class Client(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email_domain = db.Column(db.String(200), nullable=False, unique=True)
    orders = db.relationship('Order', backref='client', lazy=True)

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    management_no = db.Column(db.String(100), nullable=False)
    client_id = db.Column(db.Integer, db.ForeignKey('client.id'), nullable=False)
    filename = db.Column(db.String(200))
    amount = db.Column(db.Float, default=0)
    work_content = db.Column(db.String(500), default='')
    tantou = db.Column(db.String(100), default='')
    request_date = db.Column(db.String(50), default='')
    user_name = db.Column(db.String(200), default='')
    work_place = db.Column(db.String(300), default='')
    created_at = db.Column(db.DateTime, default=datetime.now)
    completed_at = db.Column(db.DateTime, nullable=True)

def decode_str(s):
    if s is None:
        return ''
    decoded = decode_header(s)
    result = ''
    for part, enc in decoded:
        if isinstance(part, bytes):
            result += part.decode(enc or 'utf-8', errors='ignore')
        else:
            result += part
    return result

def extract_pdf_info(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        ai_client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        message = ai_client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": """以下の依頼書テキストから情報を抽出してください。必ずJSON形式のみで返答してください。

この依頼書はアライ電機産業から合同会社LIASSへの作業依頼書です。
依頼書には「依頼会社」と「ユーザー」という2つの異なる欄があります。

抽出ルール：
- request_date: 依頼日または発行日
- work_content: 作業内容（品名・作業指示の内容）
- amount: 工賃合計金額（数字のみ）
- tantou: 「ご担当者」欄の名前のみ（「様」は除く）
- user_name: 必ず「ユーザー」と明記された欄の内容のみ。「依頼会社」欄の内容は絶対に使わない。「合同会社LIASS」「山元」「山元様」は絶対に入れない。
- work_place: 「車輌入庫場所」または「作業場所」欄の内容

重要：
- 「依頼会社」欄（例：日本カーソリューションズ、オリックス自動車、住友三井オートサービス、トヨタレンタリースなど）はuser_nameに使わない
- 「ユーザー」欄（依頼会社の下にある欄）の内容だけをuser_nameに使う
- 依頼書冒頭の「合同会社 LIASS」「山元 様」は送付先なので無視する

{"request_date": "", "work_content": "", "amount": "", "tantou": "", "user_name": "", "work_place": ""}

依頼書テキスト：
""" + text[:3000]
            }]
        )
        text_response = message.content[0].text.strip()
        text_response = text_response.replace('```json', '').replace('```', '').strip()
        return json.loads(text_response)
    except Exception as e:
        print("OCRエラー:", e)
        return {"request_date": "", "work_content": "", "amount": "", "tantou": "", "user_name": "", "work_place": ""}

def fetch_mail_pdfs(imap_server, address, password, days=90):
    results = []
    try:
        print("接続中:", address)
        mail = imaplib.IMAP4_SSL(imap_server, 993)
        mail.login(address, password)
        mail.select('inbox')
        since_date = (datetime.now() - timedelta(days=days)).strftime("%d-%b-%Y")
        _, messages = mail.search(None, 'SINCE ' + since_date)
        mail_ids = messages[0].split()
        print("メール件数:", len(mail_ids))
        for mail_id in mail_ids:
            _, msg_data = mail.fetch(mail_id, '(RFC822)')
            msg = email.message_from_bytes(msg_data[0][1])
            from_addr = decode_str(msg.get('From', ''))
            domain = from_addr.split('@')[-1].split('>')[0].strip().lower() if '@' in from_addr else ''
            for part in msg.walk():
                if part.get_content_type() == 'application/pdf':
                    filename = decode_str(part.get_filename())
                    if not filename:
                        continue
                    folder = os.path.join(app.config['UPLOAD_FOLDER'], 'mail_inbox')
                    os.makedirs(folder, exist_ok=True)
                    filepath = os.path.join(folder, filename)
                    with open(filepath, 'wb') as f:
                        f.write(part.get_payload(decode=True))
                    results.append({'filepath': filepath, 'filename': filename, 'domain': domain})
        mail.logout()
    except Exception as e:
        print("メールエラー:", e)
    return results

@app.route('/')
def index():
    clients = Client.query.all()
    orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('index.html', clients=clients, orders=orders)

@app.route('/fetch_mail', methods=['POST'])
def fetch_mail():
    all_files = []
    gmail_addr = os.getenv('GMAIL_ADDRESS', '')
    gmail_pass = os.getenv('GMAIL_APP_PASSWORD', '')
    if gmail_addr and gmail_pass:
        all_files += fetch_mail_pdfs('imap.gmail.com', gmail_addr, gmail_pass)
    outlook_addr = os.getenv('OUTLOOK_ADDRESS', '')
    outlook_pass = os.getenv('OUTLOOK_PASSWORD', '')
    if outlook_addr and outlook_pass:
        all_files += fetch_mail_pdfs('imap.goope.jp', outlook_addr, outlook_pass)
    outlook_addr2 = os.getenv('OUTLOOK_ADDRESS2', '')
    outlook_pass2 = os.getenv('OUTLOOK_PASSWORD2', '')
    if outlook_addr2 and outlook_pass2:
        all_files += fetch_mail_pdfs('imap.goope.jp', outlook_addr2, outlook_pass2)
    count = 0
    skip = 0
    for f in all_files:
        client = Client.query.filter_by(email_domain=f['domain']).first()
        if not client:
            continue
        management_no = os.path.splitext(f['filename'])[0]
        existing = Order.query.filter_by(management_no=management_no).first()
        if existing:
            skip += 1
            continue
        info = extract_pdf_info(f['filepath'])
        amount = 0
        try:
            amount = float(str(info.get('amount', '0')).replace(',', '').replace('円', '') or 0)
        except:
            amount = 0
        folder = os.path.join(app.config['UPLOAD_FOLDER'], client.name)
        os.makedirs(folder, exist_ok=True)
        dest = os.path.join(folder, f['filename'])
        try:
            os.rename(f['filepath'], dest)
        except:
            pass
        order = Order(
            management_no=management_no,
            client_id=client.id,
            filename=f['filename'],
            work_content=info.get('work_content', ''),
            amount=amount,
            tantou=info.get('tantou', ''),
            request_date=info.get('request_date', ''),
            user_name=info.get('user_name', ''),
            work_place=info.get('work_place', '')
        )
        db.session.add(order)
        db.session.commit()
        count += 1
    flash('メール取込完了：' + str(count) + '件登録、' + str(skip) + '件スキップ', 'success')
    return redirect(url_for('index'))

@app.route('/reread_all', methods=['POST'])
def reread_all():
    orders = Order.query.filter(db.or_(
        Order.user_name.like('%LIASS%'),
        Order.user_name.like('%山元%'),
        Order.user_name.like('%日本カーソリューションズ%'),
        Order.user_name.like('%オリックス自動車%'),
        Order.user_name.like('%住友三井%'),
        Order.user_name.like('%トヨタレンタリース%')
    )).all()
    count = 0
    skip = 0
    total = len(orders)
    for order in orders:
        folder = os.path.join(app.config['UPLOAD_FOLDER'], order.client.name)
        filepath = os.path.join(folder, order.filename)
        if not os.path.exists(filepath):
            print('ファイルなし:', filepath)
            skip += 1
            continue
        info = extract_pdf_info(filepath)
        if info.get('user_name'):
            order.user_name = info.get('user_name')
        if info.get('tantou'):
            order.tantou = info.get('tantou')
        if info.get('work_place'):
            order.work_place = info.get('work_place')
        db.session.commit()
        count += 1
        print('更新:', order.management_no, count, '/', total)
    flash('再読み込み完了：' + str(count) + '件更新、' + str(skip) + '件スキップ', 'success')
    return redirect(url_for('index'))

@app.route('/clients', methods=['GET', 'POST'])
def clients():
    if request.method == 'POST':
        name = request.form['name']
        email_domain = request.form['email_domain'].strip().lstrip('@')
        existing = Client.query.filter_by(email_domain=email_domain).first()
        if existing:
            flash('そのドメインは既に登録されています', 'error')
        else:
            client = Client(name=name, email_domain=email_domain)
            db.session.add(client)
            db.session.commit()
            folder = os.path.join(app.config['UPLOAD_FOLDER'], name)
            os.makedirs(folder, exist_ok=True)
            flash(name + ' を登録しました', 'success')
        return redirect(url_for('clients'))
    all_clients = Client.query.all()
    return render_template('clients.html', clients=all_clients)

@app.route('/client/<int:client_id>/orders')
def client_orders(client_id):
    client = Client.query.get_or_404(client_id)
    orders = Order.query.filter_by(client_id=client_id).order_by(Order.created_at.desc()).all()
    total = sum(o.amount or 0 for o in orders)
    now = datetime.now()
    return render_template('client_orders.html', client=client, orders=orders, total=total, now_year=now.year, now_month=now.month)

@app.route('/upload', methods=['POST'])
def upload():
    client_id = request.form.get('client_id')
    file = request.files.get('pdf')
    if not file or not client_id:
        flash('ファイルと依頼元を選択してください', 'error')
        return redirect(url_for('index'))
    client = Client.query.get(client_id)
    management_no = os.path.splitext(file.filename)[0]
    folder = os.path.join(app.config['UPLOAD_FOLDER'], client.name)
    os.makedirs(folder, exist_ok=True)
    filepath = os.path.join(folder, file.filename)
    file.save(filepath)
    info = extract_pdf_info(filepath)
    amount = 0
    try:
        amount = float(str(info.get('amount', '0')).replace(',', '').replace('円', '') or 0)
    except:
        amount = 0
    order = Order(
        management_no=management_no,
        client_id=client.id,
        filename=file.filename,
        work_content=info.get('work_content', ''),
        amount=amount,
        tantou=info.get('tantou', ''),
        request_date=info.get('request_date', ''),
        user_name=info.get('user_name', ''),
        work_place=info.get('work_place', '')
    )
    db.session.add(order)
    db.session.commit()
    flash('管理番号 ' + management_no + ' を登録しました。AI読み取り完了！', 'success')
    return redirect(url_for('index'))

@app.route('/pdf/<int:order_id>')
def view_pdf(order_id):
    order = Order.query.get_or_404(order_id)
    folder = os.path.join(app.config['UPLOAD_FOLDER'], order.client.name)
    filepath = os.path.join(folder, order.filename)
    return send_file(filepath, mimetype='application/pdf')

@app.route('/search')
def search():
    q = request.args.get('q', '')
    if q:
        orders = Order.query.filter(
            db.or_(
                Order.management_no.contains(q),
                Order.user_name.contains(q)
            )
        ).all()
    else:
        orders = []
    return render_template('search.html', orders=orders, q=q)

@app.route('/order/<int:order_id>/edit', methods=['GET', 'POST'])
def edit_order(order_id):
    order = Order.query.get_or_404(order_id)
    if request.method == 'POST':
        order.work_content = request.form.get('work_content', '')
        order.amount = float(request.form.get('amount', 0) or 0)
        order.tantou = request.form.get('tantou', '')
        order.request_date = request.form.get('request_date', '')
        order.user_name = request.form.get('user_name', '')
        order.work_place = request.form.get('work_place', '')
        db.session.commit()
        flash('管理番号 ' + order.management_no + ' を更新しました', 'success')
        return redirect(url_for('index'))
    return render_template('edit_order.html', order=order)

@app.route('/order/<int:order_id>/delete', methods=['POST'])
def delete_order(order_id):
    order = Order.query.get_or_404(order_id)
    db.session.delete(order)
    db.session.commit()
    flash('削除しました', 'success')
    return redirect(url_for('index'))

@app.route('/order/<int:order_id>/complete', methods=['GET', 'POST'])
def complete_order(order_id):
    order = Order.query.get_or_404(order_id)
    if request.method == 'POST':
        worker_name = request.form.get('worker_name', '')
        completed_date = request.form.get('completed_date', '')
        order.tantou = worker_name
        if completed_date:
            order.completed_at = datetime.strptime(completed_date, '%Y-%m-%d')
        else:
            order.completed_at = datetime.now()
        db.session.commit()
        flash('完了報告を保存しました：' + order.management_no, 'success')
        return redirect(url_for('completed'))
    return render_template('complete_order.html', order=order)

@app.route('/order/<int:order_id>/uncomplete', methods=['POST'])
def uncomplete_order(order_id):
    order = Order.query.get_or_404(order_id)
    order.completed_at = None
    db.session.commit()
    flash('完了を取り消しました：' + order.management_no, 'success')
    return redirect(url_for('index'))

@app.route('/completed')
def completed():
    orders = Order.query.filter(Order.completed_at != None).order_by(Order.completed_at.desc()).all()
    return render_template('completed.html', orders=orders)

@app.route('/export/client/<int:client_id>')
def export_client(client_id):
    import openpyxl
    client = Client.query.get_or_404(client_id)
    orders = Order.query.filter_by(client_id=client_id).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['管理番号', '依頼日', 'ユーザー名', '作業場所', '作業内容', '担当者', '完了日', '金額'])
    total = 0
    for o in orders:
        ws.append([o.management_no, o.request_date, o.user_name, o.work_place, o.work_content, o.tantou, o.completed_at.strftime('%Y/%m/%d') if o.completed_at else '', o.amount])
        total += o.amount or 0
    ws.append(['', '', '', '', '', '', '合計', total])
    path = 'invoice_' + client.name + '.xlsx'
    wb.save(path)
    return send_file(path, as_attachment=True)

@app.route('/export/invoice/<int:client_id>', methods=['POST'])
def export_invoice(client_id):
    from openpyxl import load_workbook
    client = Client.query.get_or_404(client_id)
    year = int(request.form.get('year'))
    month = int(request.form.get('month'))
    orders = Order.query.filter(
        Order.client_id == client_id,
        Order.completed_at != None,
        db.extract('year', Order.completed_at) == year,
        db.extract('month', Order.completed_at) == month
    ).order_by(Order.completed_at).all()
    if not orders:
        flash(str(year) + '年' + str(month) + '月の完了案件はありません', 'error')
        return redirect(url_for('client_orders', client_id=client_id))
    template_path = 'アライ請求書テンプレート.xlsx'
    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    ws.title = str(month) + '月'
    issue_date = str(year) + '年' + str(month) + '月末日'
    ticket_no = 'LIASS-A-' + str(year) + '-' + str(month)
    try:
        ws['M2'] = issue_date
    except:
        pass
    try:
        ws['M3'] = ticket_no
    except:
        pass
    try:
        ws['B7'] = client.name + '　御中'
    except:
        pass
    start_row = 16
    for row_num in range(start_row, ws.max_row+1):
        for col_num in range(1, 16):
            try:
                ws.cell(row=row_num, column=col_num).value = None
            except:
                pass
    subtotal = sum(o.amount or 0 for o in orders)
    tax = round(subtotal * 0.1)
    total = subtotal + tax
    try:
        ws['E12'] = total
    except:
        pass
    for row_idx, o in enumerate(orders, start_row):
        data = [
            (1, o.completed_at.strftime('%Y/%m/%d') if o.completed_at else ''),
            (4, o.management_no),
            (6, o.user_name or ''),
            (7, '様'),
            (8, '〇'),
            (9, o.work_content or ''),
            (11, o.amount or 0)
        ]
        for col, val in data:
            try:
                ws.cell(row=row_idx, column=col, value=val)
            except:
                pass
    last_row = start_row + len(orders)
    for col, val in [(1, '合計（外税）'), (6, subtotal)]:
        try:
            ws.cell(row=last_row, column=col, value=val)
        except:
            pass
    for col, val in [(1, '消費税（10%）'), (6, tax)]:
        try:
            ws.cell(row=last_row+1, column=col, value=val)
        except:
            pass
    for col, val in [(1, '合計（内税）'), (6, total)]:
        try:
            ws.cell(row=last_row+2, column=col, value=val)
        except:
            pass
    path = '請求書_' + client.name + '_' + str(year) + str(month).zfill(2) + '.xlsx'
    wb.save(path)
    return send_file(path, as_attachment=True)

@app.route('/export/all')
def export_all():
    import openpyxl
    orders = Order.query.order_by(Order.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.append(['管理番号', '依頼元', '依頼日', 'ユーザー名', '作業場所', '作業内容', '担当者', '金額', '登録日', '完了日'])
    for o in orders:
        ws.append([o.management_no, o.client.name, o.request_date, o.user_name, o.work_place, o.work_content, o.tantou, o.amount or 0, o.created_at.strftime('%Y/%m/%d'), o.completed_at.strftime('%Y/%m/%d') if o.completed_at else ''])
    path = 'all_orders.xlsx'
    wb.save(path)
    return send_file(path, as_attachment=True)

with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)